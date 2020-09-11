from nornir import InitNornir
from nornir.plugins.tasks.networking import napalm_get
from nornir.plugins.functions.text import print_result
from ttp import ttp
from openpyxl import Workbook
from openpyxl.styles import Font
import time
import re

nr = InitNornir(config_file='config.yaml')
master_data = []

def get_ints():
    get_ints_obj = nr.run(napalm_get, getters=['get_interfaces'])
    return get_ints_obj

def get_config():
    get_config_obj = nr.run(napalm_get, getters=['get_config'])
    return get_config_obj

def get_mac():
    get_mac_obj = nr.run(napalm_get, getters=['get_mac_address_table'])
    return get_mac_obj

def generate_master_data(get_ints_obj, get_config_obj, get_mac_obj):
    # Loop through hosts
    for host in nr.inventory.hosts:
        get_interfaces_host = get_ints_obj[host][0].result
        get_config_host = get_config_obj[host][0].result['get_config']['startup']
        get_mac_host = get_mac_obj[host][0].result['get_mac_address_table']
        # Parse 'show config' output against Jinja2 template
        get_config_parser = ttp(data=get_config_host, template='interface_template.j2')
        get_config_parser.parse()
        int_config = get_config_parser.result()[0][0]
        for x in get_interfaces_host:
            # Loop through both sets of data at once
            for y, z in zip(get_interfaces_host[x], int_config):
                int_info = get_interfaces_host[x][y]
                int_id = y
                # Check that interfaces are aligned in both sets of data
                if int_id != z['interface']:
                    print('Error: Interfaces are not aligned.')
                is_enabled = str(int_info['is_enabled'])
                is_up = str(int_info['is_up'])
                mode = z.get('mode', '') # If mode key is found in dictionary get value, otherwise make ''
                if mode == 'access':
                    vlan = z['access_vlan']
                    int_id_abrv = int_id[:2] + re.sub('[a-zA-Z]', '', int_id) # Abreviate interface label to be compatible with mac address table
                    # Loop through MAC address table output
                    for dictionary in get_mac_host:
                        # Check if interface is found in MAC address table
                        if dictionary['interface'] == int_id_abrv:
                            mac_address = dictionary['mac']
                            break
                        else:
                            mac_address = ''
                elif mode == 'trunk':
                    vlan = z['trunk_vlans']
                    mac_address = ''
                else:
                    vlan = ''
                    int_id_abrv = int_id[:2] + re.sub('[a-zA-Z]', '', int_id)
                    for dictionary in get_mac_host:
                        if dictionary['interface'] == int_id_abrv:
                            mac_address = dictionary['mac']
                            break
                        else:
                            mac_address = ''
                ip_address = z.get('ip_address', '')
                subnet = z.get('subnet', '')
                master_data.append([host, int_id, is_enabled, is_up, mode, vlan, mac_address, ip_address, subnet])
    return master_data

def write_to_spreadsheet(master_data):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    ws.cell(2, 2).value = 'Hostname'
    ws.cell(2, 3).value = 'Interface'
    ws.cell(2, 4).value = 'Is Enabled'
    ws.cell(2, 5).value = 'Is Up'
    ws.cell(2, 6).value = 'Mode'
    ws.cell(2, 7).value = 'VLAN(s)'
    ws.cell(2, 8).value = 'MAC Address'
    ws.cell(2, 9).value = 'IP Address'
    ws.cell(2, 10).value = 'Subnet Mask'
    ws.cell(2, 2).font = Font(bold=True)
    ws.cell(2, 3).font = Font(bold=True)
    ws.cell(2, 4).font = Font(bold=True)
    ws.cell(2, 5).font = Font(bold=True)
    ws.cell(2, 6).font = Font(bold=True)
    ws.cell(2, 7).font = Font(bold=True)
    ws.cell(2, 8).font = Font(bold=True)
    ws.cell(2, 9).font = Font(bold=True)
    ws.cell(2, 10).font = Font(bold=True)
    ROW_NUM = 3
    for interface_list in master_data:
        ws.cell(ROW_NUM, 2).value = interface_list[0]
        ws.cell(ROW_NUM, 3).value = interface_list[1]
        ws.cell(ROW_NUM, 4).value = interface_list[2]
        ws.cell(ROW_NUM, 5).value = interface_list[3]
        ws.cell(ROW_NUM, 6).value = interface_list[4]
        ws.cell(ROW_NUM, 7).value = interface_list[5]
        ws.cell(ROW_NUM, 8).value = interface_list[6]
        ws.cell(ROW_NUM, 9).value = interface_list[7]
        ws.cell(ROW_NUM, 10).value = interface_list[8]
        ROW_NUM += 1
    timestr = time.strftime('%Y%m%d_%H%M%S')
    wb.save('Network_Survey_' + timestr + '.xlsx')

if __name__ == '__main__':
    get_ints_obj = get_ints()
    get_config_obj = get_config()
    get_mac_obj = get_mac()
    master_data = generate_master_data(get_ints_obj, get_config_obj, get_mac_obj)
    write_to_spreadsheet(master_data)